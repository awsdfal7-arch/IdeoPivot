from __future__ import annotations

from datetime import date
from decimal import Decimal
import json
import threading
import time

from openpyxl import Workbook, load_workbook
import pytest
from sj_generator.infrastructure.persistence.sqlite_repo import DbQuestionRecord, append_questions, delete_question_by_id, load_all_questions

from sj_generator.infrastructure.llm import balance as balance_mod
from sj_generator.infrastructure.llm import explanations as explanations_mod
from sj_generator.infrastructure.llm import import_questions as iq
from sj_generator.infrastructure.llm import prompt_templates as prompt_mod
from sj_generator.infrastructure.llm.client import _pick_temperature
from sj_generator.infrastructure.llm.import_questions import ImportResult
from sj_generator.infrastructure.llm.task_runner import run_callables_in_parallel_fail_fast, run_tasks_in_parallel
from sj_generator.application.importing import batch_ai_import as batch_ai
from sj_generator.application.exporting.batch_folderize import process_excel_to_folder_mode
from sj_generator.infrastructure.persistence.draft_db_import import draft_questions_to_db_records
from sj_generator.infrastructure.exporting.export_md import _normalize_numbers, export_questions_to_markdown
from sj_generator.domain.entities import Question
from sj_generator.application import settings as cfg_mod
from sj_generator.application.settings.import_cost_history import (
    append_balance_history_for_provider_results,
    clear_import_cost_history,
    append_import_cost_history_entry,
    build_total_balance_text,
    load_import_cost_history_rows,
)
from sj_generator.ui.compare_highlight import compare_highlight_model_styles
from sj_generator.ui.import_costs import (
    _reset_app_import_cost_capture_for_tests,
    begin_app_import_cost_capture,
    capture_import_cost_before,
    capture_import_cost_before_async,
    freeze_import_cost_result,
    wait_import_cost_before_capture,
)
from sj_generator.application.state import (
    ImportWizardSession,
    WizardState,
    build_import_flow_session,
    default_repo_parent_dir,
    normalize_default_repo_parent_dir_text,
    normalize_ai_concurrency,
    normalize_analysis_model_name,
    normalize_analysis_provider,
    normalize_export_include_answers,
    normalize_export_include_analysis,
)


def test_normalize_numbers_fill_after_existing_max() -> None:
    questions = [
        Question(number="5", stem="s1", options="", answer="", analysis=""),
        Question(number="", stem="s2", options="", answer="", analysis=""),
        Question(number="7", stem="s3", options="", answer="", analysis=""),
        Question(number="", stem="s4", options="", answer="", analysis=""),
    ]
    normalized = _normalize_numbers(questions)
    assert [q.number for q in normalized] == ["5", "8", "7", "9"]


def test_delete_question_by_id_removes_only_target_record(tmp_path) -> None:
    db_path = tmp_path / "questions.db"
    records = [
        DbQuestionRecord(
            id="q1",
            stem="题目一",
            option_1="甲",
            option_2="乙",
            option_3="丙",
            option_4="丁",
            choice_1="",
            choice_2="",
            choice_3="",
            choice_4="",
            answer="A",
            analysis="解析一",
            question_type="单选",
            textbook_version="必修一",
            source="资料一.docx",
            level_path="1.1",
            difficulty_score=None,
            knowledge_points="",
            abilities="",
            created_at="2026-04-21 10:00:00",
            updated_at="2026-04-21 10:00:00",
        ),
        DbQuestionRecord(
            id="q2",
            stem="题目二",
            option_1="甲",
            option_2="乙",
            option_3="丙",
            option_4="丁",
            choice_1="",
            choice_2="",
            choice_3="",
            choice_4="",
            answer="B",
            analysis="解析二",
            question_type="单选",
            textbook_version="必修一",
            source="资料二.docx",
            level_path="1.1",
            difficulty_score=None,
            knowledge_points="",
            abilities="",
            created_at="2026-04-21 10:01:00",
            updated_at="2026-04-21 10:01:00",
        ),
    ]
    append_questions(db_path, records)

    assert delete_question_by_id(db_path, "q1") == 1
    assert [record.id for record in load_all_questions(db_path)] == ["q2"]


def test_config_path_defaults_to_appdata(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_CONFIG_PATH", raising=False)
    monkeypatch.delenv("SJ_GENERATOR_KIMI_CONFIG_PATH", raising=False)
    monkeypatch.delenv("SJ_GENERATOR_QWEN_CONFIG_PATH", raising=False)
    monkeypatch.delenv("SJ_GENERATOR_PROGRAM_SETTINGS_PATH", raising=False)

    assert cfg_mod._config_path() == tmp_path / "sj_generator" / "deepseek.json"
    assert cfg_mod._kimi_config_path() == tmp_path / "sj_generator" / "kimi.json"
    assert cfg_mod._qwen_config_path() == tmp_path / "sj_generator" / "qwen.json"
    assert cfg_mod._program_settings_path() == tmp_path / "sj_generator" / "program_settings.json"


def test_program_settings_persist_round_trip(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_PROGRAM_SETTINGS_PATH", raising=False)

    cfg_mod.save_program_settings(
        {
            "default_repo_parent_dir_text": "C:/repo-root",
            "ai_concurrency": 4,
            "analysis_enabled": False,
            "import_show_costs": False,
            "dedupe_enabled": True,
            "analysis_provider": "kimi",
            "analysis_model_name": "kimi-k2-turbo-preview",
            "export_convertible_multi_mode": "as_multi",
            "export_include_answers": False,
            "export_include_analysis": True,
            "preferred_textbook_version": "必修二",
        }
    )

    assert cfg_mod.load_program_settings() == {
        "default_repo_parent_dir_text": "C:/repo-root",
        "ai_concurrency": 4,
        "analysis_enabled": False,
        "import_show_costs": False,
        "dedupe_enabled": True,
        "analysis_provider": "kimi",
        "analysis_model_name": "kimi-k2-turbo-preview",
        "export_convertible_multi_mode": "as_multi",
        "export_include_answers": False,
        "export_include_analysis": True,
        "preferred_textbook_version": "必修二",
    }


def test_save_program_settings_merged_preserves_existing_values(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_PROGRAM_SETTINGS_PATH", raising=False)

    cfg_mod.save_program_settings(
        {
            "analysis_provider": "deepseek",
            "analysis_model_name": "deepseek-reasoner",
            "project_parse_model_rows": [
                {
                    "key": "question_number_parse",
                    "round": "1",
                    "ratio": "1/4",
                    "models": [{"provider": "deepseek", "model_name": "deepseek-chat"}],
                },
                {
                    "key": "question_content_parse",
                    "round": "2",
                    "ratio": "1/4",
                    "models": [{"provider": "kimi", "model_name": "kimi-k2.6"}],
                },
            ],
        }
    )

    cfg_mod.save_program_settings_merged(
        {
            "question_content_concurrency": 4,
            "analysis_generation_concurrency": 2,
        }
    )

    saved = cfg_mod.load_program_settings()
    assert saved["analysis_provider"] == "deepseek"
    assert saved["analysis_model_name"] == "deepseek-reasoner"
    assert saved["question_content_concurrency"] == 4
    assert saved["analysis_generation_concurrency"] == 2
    assert saved["project_parse_model_rows"][0]["models"][0] == {
        "provider": "deepseek",
        "model_name": "deepseek-chat",
    }


def test_save_program_analysis_target_persists(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_PROGRAM_SETTINGS_PATH", raising=False)
    cfg_mod.save_program_settings({"analysis_provider": "deepseek", "analysis_model_name": "deepseek-reasoner"})

    cfg_mod.save_program_analysis_target(provider="kimi", model_name="kimi-k2.6")

    saved = cfg_mod.load_program_settings()
    assert saved["analysis_provider"] == "kimi"
    assert saved["analysis_model_name"] == "kimi-k2.6"


def test_import_cost_tracking_builds_summary_from_balance_delta(monkeypatch) -> None:
    _reset_app_import_cost_capture_for_tests()
    state = ImportWizardSession(import_show_costs=True)

    class Snapshot:
        def __init__(self, provider: str, currency: str, amount: str, detail: str) -> None:
            self.provider = provider
            self.currency = currency
            self.amount = Decimal(amount)
            self.detail = detail

    calls = {"count": 0}

    def fake_query_provider_balance_snapshots(**_kwargs):
        calls["count"] += 1
        if calls["count"] == 1:
            return [
                Snapshot("deepseek", "CNY", "100", "已配置，余额 CNY ¥100.00"),
                Snapshot("kimi", "CNY", "50", "已配置，余额 CNY ¥50.00"),
            ]
        return [
            Snapshot("deepseek", "CNY", "98.5", "已配置，余额 CNY ¥98.50"),
            Snapshot("kimi", "CNY", "50", "已配置，余额 CNY ¥50.00"),
        ]

    monkeypatch.setattr("sj_generator.ui.import_costs.query_provider_balance_snapshots", fake_query_provider_balance_snapshots)
    capture_import_cost_before(state)
    freeze_import_cost_result(state)

    assert state.execution.import_cost_ready is True
    assert state.execution.import_cost_total_text == "¥1.5000"
    assert "DeepSeek" in state.execution.import_cost_summary_text
    assert "¥1.5000" in state.execution.import_cost_summary_text
    assert "Kimi" in state.execution.import_cost_detail_text
    assert any(row.get("provider") == "Kimi" for row in state.execution.import_cost_rows)


def test_import_cost_before_async_does_not_block_page_transition(monkeypatch) -> None:
    _reset_app_import_cost_capture_for_tests()
    state = ImportWizardSession(import_show_costs=True)
    started = threading.Event()
    release = threading.Event()

    class Snapshot:
        def __init__(self, provider: str, currency: str, amount: str, detail: str) -> None:
            self.provider = provider
            self.currency = currency
            self.amount = Decimal(amount)
            self.detail = detail

    def fake_query_provider_balance_snapshots(**_kwargs):
        started.set()
        release.wait(2)
        return [Snapshot("deepseek", "CNY", "100", "已配置，余额 CNY ¥100.00")]

    monkeypatch.setattr("sj_generator.ui.import_costs.query_provider_balance_snapshots", fake_query_provider_balance_snapshots)

    capture_import_cost_before_async(state)

    assert started.wait(1) is True
    assert state.execution.import_cost_before_loading is True
    assert state.execution.import_cost_before_amounts == {}

    release.set()
    freeze_import_cost_result(state)

    assert state.execution.import_cost_before_loading is False
    assert state.execution.import_cost_before_amounts["deepseek"] == "CNY|100"


def test_app_import_cost_capture_retries_when_initial_result_is_empty(monkeypatch) -> None:
    _reset_app_import_cost_capture_for_tests()
    state = ImportWizardSession(import_show_costs=True)

    class Snapshot:
        def __init__(self, provider: str, currency: str, amount: str, detail: str) -> None:
            self.provider = provider
            self.currency = currency
            self.amount = Decimal(amount)
            self.detail = detail

    calls = {"count": 0}

    def fake_query_provider_balance_snapshots(**_kwargs):
        calls["count"] += 1
        if calls["count"] == 1:
            return []
        return [Snapshot("deepseek", "CNY", "100", "已配置，余额 CNY ¥100.00")]

    monkeypatch.setattr("sj_generator.ui.import_costs.query_provider_balance_snapshots", fake_query_provider_balance_snapshots)

    begin_app_import_cost_capture()
    wait_import_cost_before_capture(state)
    assert state.execution.import_cost_before_amounts == {}

    begin_app_import_cost_capture(retry_if_empty=True)
    wait_import_cost_before_capture(state)

    assert calls["count"] == 2
    assert state.execution.import_cost_before_amounts["deepseek"] == "CNY|100"


def test_app_import_cost_capture_writes_startup_balance_log(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("SJ_GENERATOR_IMPORT_COST_HISTORY_PATH", str(tmp_path / "import_cost_history.json"))
    _reset_app_import_cost_capture_for_tests()
    state = ImportWizardSession(import_show_costs=True)

    class Snapshot:
        def __init__(self, provider: str, currency: str, amount: str, detail: str) -> None:
            self.provider = provider
            self.currency = currency
            self.amount = Decimal(amount)
            self.detail = detail

    def fake_query_provider_balance_snapshots(**_kwargs):
        return [
            Snapshot("deepseek", "CNY", "100", "已配置，余额 CNY ¥100.00"),
            Snapshot("kimi", "CNY", "50", "已配置，余额 CNY ¥50.00"),
        ]

    monkeypatch.setattr("sj_generator.ui.import_costs.query_provider_balance_snapshots", fake_query_provider_balance_snapshots)

    begin_app_import_cost_capture()
    wait_import_cost_before_capture(state)

    rows = load_import_cost_history_rows()

    assert len(rows) == 1
    assert rows[0]["run_at"].endswith("[启动]")
    assert rows[0]["deepseek_balance"] == "¥100.0000"
    assert rows[0]["kimi_balance"] == "¥50.0000"
    assert rows[0]["qwen_balance"] == ""


def test_test_all_results_append_balance_history(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("SJ_GENERATOR_IMPORT_COST_HISTORY_PATH", str(tmp_path / "import_cost_history.json"))

    append_balance_history_for_provider_results(
        [
            ("deepseek", {"balance_value": "¥10.5000"}),
            ("kimi", {"balance_value": "¥20.0000"}),
            ("qwen", {"balance_value": "¥30.0000"}),
        ],
        source_label="统一测试",
    )

    rows = load_import_cost_history_rows()

    assert len(rows) == 1
    assert rows[0]["run_at"].endswith("[统一测试]")
    assert rows[0]["deepseek_balance"] == "¥10.5000"
    assert rows[0]["kimi_balance"] == "¥20.0000"
    assert rows[0]["qwen_balance"] == "¥30.0000"


def test_clear_import_cost_history_empties_rows(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("SJ_GENERATOR_IMPORT_COST_HISTORY_PATH", str(tmp_path / "import_cost_history.json"))

    append_import_cost_history_entry(
        run_at="2026-04-26 12:00:00",
        provider_balances={"deepseek": "¥10.0000", "kimi": "¥20.0000", "qwen": "¥30.0000"},
    )

    clear_import_cost_history()

    assert load_import_cost_history_rows() == []


def test_import_cost_history_round_trip(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("SJ_GENERATOR_IMPORT_COST_HISTORY_PATH", str(tmp_path / "import_cost_history.json"))

    append_import_cost_history_entry(
        run_at="2026-04-26 10:00:00",
        before_provider_balances={"deepseek": "¥100.0000", "kimi": "¥50.0000", "qwen": ""},
        provider_balances={"deepseek": "¥98.5000", "kimi": "¥50.0000", "qwen": ""},
        total_balance="¥148.5000",
        total_cost="¥1.5000",
        cost_summary="DeepSeek ¥1.5000",
        cost_detail="DeepSeek：¥1.5000",
    )
    append_import_cost_history_entry(
        run_at="2026-04-26 11:00:00",
        before_provider_balances={"deepseek": "¥98.5000", "kimi": "¥50.0000", "qwen": "$20.5000"},
        provider_balances={"deepseek": "¥97.0000", "kimi": "¥49.5000", "qwen": "$20.0000"},
        total_balance="¥146.5000；$20.0000",
        total_cost="¥2.0000",
        cost_summary="DeepSeek ¥1.5000；Kimi ¥0.5000",
        cost_detail="DeepSeek：¥1.5000；Kimi：¥0.5000",
    )

    rows = load_import_cost_history_rows()

    assert [row["run_at"] for row in rows] == ["2026-04-26 11:00:00", "2026-04-26 10:00:00"]
    assert rows[0]["deepseek_balance"] == "¥97.0000"
    assert rows[0]["kimi_balance"] == "¥49.5000"
    assert rows[0]["qwen_balance"] == "$20.0000"


def test_import_cost_history_prefers_more_complete_same_second_row(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("SJ_GENERATOR_IMPORT_COST_HISTORY_PATH", str(tmp_path / "import_cost_history.json"))

    append_import_cost_history_entry(
        run_at="2026-04-26 10:00:00",
        before_provider_balances={"deepseek": "¥100.0000", "kimi": "", "qwen": ""},
        provider_balances={"deepseek": "¥100.0000", "kimi": "", "qwen": ""},
        total_balance="¥100.0000",
        total_cost="¥0.0000",
        cost_summary="0",
        cost_detail="DeepSeek：¥0.0000",
    )
    append_import_cost_history_entry(
        run_at="2026-04-26 10:00:00",
        before_provider_balances={"deepseek": "¥100.0000", "kimi": "¥50.0000", "qwen": ""},
        provider_balances={"deepseek": "¥98.5000", "kimi": "¥50.0000", "qwen": ""},
        total_balance="¥148.5000",
        total_cost="¥1.5000",
        cost_summary="DeepSeek ¥1.5000",
        cost_detail="DeepSeek：¥1.5000；Kimi：¥0.0000",
    )

    rows = load_import_cost_history_rows()

    assert len(rows) == 1
    assert rows[0]["deepseek_balance"] == "¥98.5000"
    assert rows[0]["kimi_balance"] == "¥50.0000"
    assert rows[0]["qwen_balance"] == ""


def test_build_total_balance_text_sums_by_currency() -> None:
    total_text = build_total_balance_text(
        {
            "deepseek": "¥98.5000",
            "kimi": "¥49.5000",
            "qwen": "$20.0000",
        }
    )

    assert total_text == "¥148.0000；$20.0000"


def test_freeze_import_cost_result_appends_history_log(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("SJ_GENERATOR_IMPORT_COST_HISTORY_PATH", str(tmp_path / "import_cost_history.json"))
    _reset_app_import_cost_capture_for_tests()
    state = ImportWizardSession(import_show_costs=True)

    class Snapshot:
        def __init__(self, provider: str, currency: str, amount: str, detail: str) -> None:
            self.provider = provider
            self.currency = currency
            self.amount = Decimal(amount)
            self.detail = detail

    calls = {"count": 0}

    def fake_query_provider_balance_snapshots(**_kwargs):
        calls["count"] += 1
        if calls["count"] == 1:
            return [
                Snapshot("deepseek", "CNY", "100", "已配置，余额 CNY ¥100.00"),
                Snapshot("kimi", "CNY", "60", "已配置，余额 CNY ¥60.00"),
            ]
        return [
            Snapshot("deepseek", "CNY", "99", "已配置，余额 CNY ¥99.00"),
            Snapshot("kimi", "CNY", "59.5", "已配置，余额 CNY ¥59.50"),
        ]

    monkeypatch.setattr("sj_generator.ui.import_costs.query_provider_balance_snapshots", fake_query_provider_balance_snapshots)

    capture_import_cost_before(state)
    freeze_import_cost_result(state)

    rows = load_import_cost_history_rows()

    assert len(rows) == 2
    assert rows[0]["run_at"].endswith("[导入完成]")
    assert rows[0]["deepseek_balance"] == "¥99.0000"
    assert rows[0]["kimi_balance"] == "¥59.5000"
    assert rows[0]["qwen_balance"] == ""
    assert state.execution.import_cost_total_text == "¥1.5000"


def test_import_cost_tracking_rolls_forward_after_each_import(monkeypatch) -> None:
    _reset_app_import_cost_capture_for_tests()
    first_state = ImportWizardSession(import_show_costs=True)
    second_state = ImportWizardSession(import_show_costs=True)

    class Snapshot:
        def __init__(self, provider: str, currency: str, amount: str, detail: str) -> None:
            self.provider = provider
            self.currency = currency
            self.amount = Decimal(amount)
            self.detail = detail

    calls = {"count": 0}

    def fake_query_provider_balance_snapshots(**_kwargs):
        calls["count"] += 1
        if calls["count"] == 1:
            return [Snapshot("deepseek", "CNY", "100", "已配置，余额 CNY ¥100.00")]
        if calls["count"] == 2:
            return [Snapshot("deepseek", "CNY", "98.5", "已配置，余额 CNY ¥98.50")]
        return [Snapshot("deepseek", "CNY", "98.0", "已配置，余额 CNY ¥98.00")]

    monkeypatch.setattr("sj_generator.ui.import_costs.query_provider_balance_snapshots", fake_query_provider_balance_snapshots)

    freeze_import_cost_result(first_state)
    freeze_import_cost_result(second_state)

    assert first_state.execution.import_cost_total_text == "¥1.5000"
    assert second_state.execution.import_cost_total_text == "¥0.5000"




def test_build_import_flow_session_copies_settings_and_source_inputs(tmp_path) -> None:
    base_state = WizardState(
        project_dir=tmp_path / "导入项目",
        repo_path=tmp_path / "导入项目" / "导入项目.xlsx",
        last_export_dir=tmp_path / "导出",
        project_name_is_placeholder=True,
        dedupe_enabled=False,
        analysis_enabled=False,
        preferred_textbook_version="必修二",
    )
    source_path = tmp_path / "资料一.docx"
    session = build_import_flow_session(base_state, source_files=[source_path])

    assert session.project_dir == base_state.project_dir
    assert session.repo_path == base_state.repo_path
    assert session.last_export_dir == base_state.last_export_dir
    assert session.project_name_is_placeholder is True
    assert session.dedupe_enabled is False
    assert session.analysis_enabled is False
    assert session.source.files == [source_path]
    assert session.source.files_text == str(source_path)
    assert session.source.file_items[0].version == "必修二"


def test_import_session_apply_question_refs_resets_downstream_state() -> None:
    session = ImportWizardSession()
    session.apply_draft_questions([Question(number="1", stem="题干", options="", answer="A", analysis="")])
    session.set_dedupe_hits([])
    session.execution.mark_db_import_completed(3)

    session.apply_question_refs({"a.docx": [{"number": "1", "question_type": "单选"}]})

    assert session.refs.revision == 1
    assert session.refs.question_refs_by_source["a.docx"][0]["number"] == "1"
    assert session.draft.questions == []
    assert session.draft.dedupe_hits is None
    assert session.execution.db_import_completed is False


def test_import_session_build_import_session_preserves_settings_for_child_flow(tmp_path) -> None:
    session = ImportWizardSession(
        project_dir=tmp_path / "项目A",
        repo_path=tmp_path / "项目A" / "项目A.xlsx",
        analysis_enabled=False,
        preferred_textbook_version="选择性必修一",
    )

    child = session.build_import_session(source_files=[tmp_path / "资料二.docx"])

    assert child.project_dir == session.project_dir
    assert child.repo_path == session.repo_path
    assert child.analysis_enabled is False
    assert child.preferred_textbook_version == "选择性必修一"
    assert child.source.files_text.endswith("资料二.docx")


def test_project_parse_model_rows_preserve_new_concurrency_settings(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_PROGRAM_SETTINGS_PATH", raising=False)
    cfg_mod.save_program_settings(
        {
            "question_content_concurrency": 5,
            "analysis_generation_concurrency": 2,
        }
    )

    cfg_mod.save_project_parse_model_rows(
        [
            {
                "key": "question_number_parse",
                "round": "1",
                "ratio": "2/3",
                "models": [{"provider": "deepseek", "model_name": "deepseek-chat"}],
            },
            {
                "key": "question_content_parse",
                "round": "2",
                "ratio": "1/4",
                "models": [{"provider": "qwen", "model_name": "qwen-max"}],
            },
        ]
    )

    saved = cfg_mod.load_program_settings()
    assert saved["question_content_concurrency"] == 5
    assert saved["analysis_generation_concurrency"] == 2
    assert saved["project_parse_model_rows"][1]["models"][0] == {"provider": "qwen", "model_name": "qwen-max"}


def test_project_parse_model_rows_persist(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_PROGRAM_SETTINGS_PATH", raising=False)

    cfg_mod.save_project_parse_model_rows(
        [
            {
                "key": "question_number_parse",
                "round": "1",
                "ratio": "2/3",
                "models": [
                    {"provider": "deepseek", "model_name": "deepseek-chat"},
                    {"provider": "kimi", "model_name": "kimi-k2.6"},
                ],
            },
            {
                "key": "question_content_parse",
                "round": "2",
                "ratio": "1/4",
                "models": [{"provider": "qwen", "model_name": "qwen-max"}],
            },
        ]
    )

    rows = cfg_mod.load_project_parse_model_rows()
    assert rows[0]["key"] == "question_number_parse"
    assert rows[0]["round"] == "1"
    assert rows[0]["ratio"] == "2/3"
    assert rows[0]["models"][0] == {"provider": "deepseek", "model_name": "deepseek-chat"}
    assert rows[1]["key"] == "question_content_parse"
    assert rows[1]["models"][0] == {"provider": "qwen", "model_name": "qwen-max"}


def test_question_number_round_is_always_fixed_to_one(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_PROGRAM_SETTINGS_PATH", raising=False)

    cfg_mod.save_project_parse_model_rows(
        [
            {
                "key": "question_number_parse",
                "round": "8",
                "ratio": "1/4",
                "models": [],
            },
            {
                "key": "question_content_parse",
                "round": "3",
                "ratio": "1/4",
                "models": [],
            },
        ]
    )

    rows = cfg_mod.load_project_parse_model_rows()
    assert rows[0]["round"] == "1"
    assert rows[1]["round"] == "3"


def test_question_content_round_limit_respects_saved_project_config(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_PROGRAM_SETTINGS_PATH", raising=False)

    cfg_mod.save_project_parse_model_rows(
        [
            {
                "key": "question_number_parse",
                "round": "1",
                "ratio": "1/4",
                "models": [],
            },
            {
                "key": "question_content_parse",
                "round": "8",
                "ratio": "1/4",
                "models": [],
            },
        ]
    )

    assert cfg_mod.load_project_parse_model_rows()[1]["round"] == "8"
    assert iq.question_content_round_limit() == 8


def test_load_deepseek_config_prefers_project_parse_model_rows(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_PROGRAM_SETTINGS_PATH", raising=False)
    monkeypatch.delenv("SJ_GENERATOR_CONFIG_PATH", raising=False)
    monkeypatch.delenv("DEEPSEEK_QUESTION_NUMBER_MODEL", raising=False)
    monkeypatch.delenv("DEEPSEEK_QUESTION_UNIT_MODEL", raising=False)
    monkeypatch.delenv("DEEPSEEK_MODEL", raising=False)

    cfg_mod.save_deepseek_config(
        cfg_mod.DeepSeekConfig(
            base_url="https://api.deepseek.com",
            api_key="",
            number_model="file-number-model",
            model="file-unit-model",
            analysis_model="deepseek-reasoner",
            timeout_s=120.0,
        )
    )
    cfg_mod.save_project_parse_model_rows(
        [
            {
                "key": "question_number_parse",
                "round": "1",
                "ratio": "1/3",
                "models": [{"provider": "deepseek", "model_name": "project-number-model"}],
            },
            {
                "key": "question_content_parse",
                "round": "2",
                "ratio": "1/3",
                "models": [{"provider": "deepseek", "model_name": "project-unit-model"}],
            },
        ]
    )

    cfg = cfg_mod.load_deepseek_config()
    assert cfg.number_model == "project-number-model"
    assert cfg.model == "project-unit-model"


def test_normalize_export_include_defaults_true() -> None:
    assert normalize_export_include_answers(None) is True
    assert normalize_export_include_answers(False) is False
    assert normalize_export_include_analysis(None) is True
    assert normalize_export_include_analysis(False) is False


def test_welcome_table_font_point_size_persists_round_trip(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    cfg_mod.save_welcome_table_font_point_size(15)
    assert cfg_mod.load_welcome_table_font_point_size() == 15


def test_draft_questions_to_db_records_applies_textbook_version_preference() -> None:
    records = draft_questions_to_db_records(
        questions=[Question(number="1", stem="题目", options="A.甲\nB.乙", answer="A", analysis="")],
        level_path="1.1.1",
        textbook_version="必修三",
    )

    assert len(records) == 1
    assert records[0].textbook_version == "必修三"


def test_export_questions_to_markdown_can_skip_answers_and_analysis() -> None:
    markdown = export_questions_to_markdown(
        excel_file_name="必修三",
        export_date=date(2026, 4, 22),
        questions=[
            Question(
                number="1",
                stem="测试题目",
                options="A. 选项一\nB. 选项二",
                answer="A",
                analysis="测试解析",
            )
        ],
        include_answers=False,
        include_analysis=False,
    )

    assert "## 答案与解析" not in markdown
    assert "测试解析" not in markdown
    assert " 1. 测试题目（ ）" in markdown


def test_export_questions_to_markdown_can_export_analysis_only() -> None:
    markdown = export_questions_to_markdown(
        excel_file_name="必修三",
        export_date=date(2026, 4, 22),
        questions=[
            Question(
                number="1",
                stem="测试题目",
                options="A. 选项一\nB. 选项二",
                answer="A",
                analysis="测试解析",
            )
        ],
        include_answers=False,
        include_analysis=True,
    )

    assert "## 解析" in markdown
    assert "测试解析" in markdown
    assert "**1**" in markdown
    assert "**1. A**" not in markdown


def test_deepseek_analysis_model_defaults_and_persists(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_CONFIG_PATH", raising=False)
    monkeypatch.delenv("DEEPSEEK_API_KEY", raising=False)
    monkeypatch.delenv("DEEPSEEK_ANALYSIS_MODEL", raising=False)
    monkeypatch.delenv("DEEPSEEK_QUESTION_NUMBER_MODEL", raising=False)

    cfg = cfg_mod.load_deepseek_config()
    assert cfg.analysis_model == "deepseek-reasoner"
    assert cfg.number_model == "deepseek-chat"

    updated = cfg_mod.DeepSeekConfig(
        base_url=cfg.base_url,
        api_key="sk-test",
        number_model="deepseek-chat",
        model="deepseek-chat",
        analysis_model="deepseek-reasoner",
        timeout_s=cfg.timeout_s,
    )
    cfg_mod.save_deepseek_config(updated)
    reloaded = cfg_mod.load_deepseek_config()
    assert reloaded.analysis_model == "deepseek-reasoner"
    assert reloaded.number_model == "deepseek-chat"
    assert reloaded.api_key == ""


def test_deepseek_api_key_only_reads_from_env_and_is_not_saved(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_CONFIG_PATH", raising=False)
    monkeypatch.delenv("DEEPSEEK_API_KEY", raising=False)

    cfg = cfg_mod.DeepSeekConfig(
        base_url="https://api.deepseek.com",
        api_key="sk-file",
        model="deepseek-chat",
        analysis_model="deepseek-reasoner",
        timeout_s=60.0,
    )
    cfg_mod.save_deepseek_config(cfg)

    saved = cfg_mod._read_json_dict(cfg_mod._config_path())
    assert "api_key" not in saved
    assert cfg_mod.load_deepseek_config().api_key == ""

    monkeypatch.setenv("DEEPSEEK_API_KEY", "sk-env")
    assert cfg_mod.load_deepseek_config().api_key == "sk-env"


def test_qwen_credentials_saved_to_file_and_file_overrides_env(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_QWEN_CONFIG_PATH", raising=False)
    isolated_settings_path = tmp_path / "program_settings.json"
    isolated_settings_path.write_text("{}", encoding="utf-8")
    monkeypatch.setenv("SJ_GENERATOR_PROGRAM_SETTINGS_PATH", str(isolated_settings_path))
    monkeypatch.delenv("QWEN_BASE_URL", raising=False)
    monkeypatch.delenv("QWEN_API_KEY", raising=False)
    monkeypatch.delenv("QWEN_MODEL", raising=False)
    monkeypatch.delenv("QWEN_QUESTION_NUMBER_MODEL", raising=False)
    monkeypatch.delenv("QWEN_QUESTION_UNIT_MODEL", raising=False)
    monkeypatch.delenv("QWEN_TIMEOUT_S", raising=False)
    monkeypatch.delenv("QWEN_ACCOUNT_ACCESS_KEY_ID", raising=False)
    monkeypatch.delenv("QWEN_ACCOUNT_ACCESS_KEY_SECRET", raising=False)
    monkeypatch.delenv("ALIBABA_CLOUD_ACCESS_KEY_ID", raising=False)
    monkeypatch.delenv("ALIBABA_CLOUD_ACCESS_KEY_SECRET", raising=False)

    cfg = cfg_mod.QwenConfig(
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
        api_key="dashscope-key",
        number_model="qwen-max",
        model="qwen-max",
        account_access_key_id="akid",
        account_access_key_secret="aksecret",
        timeout_s=60.0,
    )
    cfg_mod.save_qwen_config(cfg)

    saved = cfg_mod._read_json_dict(cfg_mod._qwen_config_path())
    assert saved["api_key"] == "dashscope-key"
    assert saved["account_access_key_id"] == "akid"
    assert saved["account_access_key_secret"] == "aksecret"

    reloaded = cfg_mod.load_qwen_config()
    assert reloaded.api_key == "dashscope-key"
    assert reloaded.number_model == "qwen-max"
    assert reloaded.account_access_key_id == "akid"
    assert reloaded.account_access_key_secret == "aksecret"
    assert reloaded.has_account_balance_credentials() is True

    monkeypatch.setenv("QWEN_API_KEY", "dashscope-env")
    monkeypatch.setenv("QWEN_ACCOUNT_ACCESS_KEY_ID", "akid-env")
    monkeypatch.setenv("QWEN_ACCOUNT_ACCESS_KEY_SECRET", "aksecret-env")
    monkeypatch.setenv("QWEN_BASE_URL", "https://dashscope.example/v1")
    monkeypatch.setenv("QWEN_QUESTION_NUMBER_MODEL", "qwen-number-model")
    monkeypatch.setenv("QWEN_QUESTION_UNIT_MODEL", "qwen-unit-model")
    monkeypatch.setenv("QWEN_TIMEOUT_S", "88")
    reloaded = cfg_mod.load_qwen_config()
    assert reloaded.base_url == "https://dashscope.aliyuncs.com/compatible-mode/v1"
    assert reloaded.api_key == "dashscope-key"
    assert reloaded.number_model == "qwen-max"
    assert reloaded.model == "qwen-max"
    assert reloaded.account_access_key_id == "akid"
    assert reloaded.account_access_key_secret == "aksecret"
    assert reloaded.timeout_s == 60.0
    assert reloaded.has_account_balance_credentials() is True


def test_qwen_available_models_include_qwen36_plus(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_QWEN_CONFIG_PATH", raising=False)

    models = cfg_mod.load_available_models("qwen")

    assert "qwen3.6-plus" in models


def test_save_qwen_config_preserves_available_models(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_QWEN_CONFIG_PATH", raising=False)

    saved_models = cfg_mod.save_available_models("qwen", ["qwen-max", "qwen3.6-plus", "qwen-plus"])
    assert saved_models == ["qwen-max", "qwen3.6-plus", "qwen-plus"]

    cfg = cfg_mod.QwenConfig(
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
        api_key="dashscope-key",
        number_model="qwen-max",
        model="qwen-max",
        account_access_key_id="",
        account_access_key_secret="",
        timeout_s=60.0,
    )
    cfg_mod.save_qwen_config(cfg)

    saved = cfg_mod._read_json_dict(cfg_mod._qwen_config_path())
    assert saved["available_models"] == ["qwen-max", "qwen3.6-plus", "qwen-plus"]


def test_kimi_question_number_model_falls_back_to_legacy_model(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_KIMI_CONFIG_PATH", raising=False)
    monkeypatch.delenv("KIMI_BASE_URL", raising=False)
    monkeypatch.delenv("KIMI_API_KEY", raising=False)
    monkeypatch.delenv("KIMI_QUESTION_NUMBER_MODEL", raising=False)
    monkeypatch.delenv("KIMI_QUESTION_UNIT_MODEL", raising=False)
    monkeypatch.delenv("KIMI_MODEL", raising=False)
    monkeypatch.delenv("KIMI_TIMEOUT_S", raising=False)
    path = cfg_mod._kimi_config_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {
                "base_url": "https://api.moonshot.cn/v1",
                "model": "kimi-k2-turbo-preview",
                "timeout_s": 60.0,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    cfg = cfg_mod.load_kimi_config()

    assert cfg.number_model == "kimi-k2-turbo-preview"
    assert cfg.model == "kimi-k2-turbo-preview"


def test_kimi_defaults_now_use_kimi_2_6(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    cfg_path = tmp_path / "isolated-kimi-defaults.json"
    cfg_path.write_text("{}", encoding="utf-8")
    monkeypatch.setenv("SJ_GENERATOR_KIMI_CONFIG_PATH", str(cfg_path))
    monkeypatch.delenv("KIMI_BASE_URL", raising=False)
    monkeypatch.delenv("KIMI_API_KEY", raising=False)
    monkeypatch.delenv("KIMI_QUESTION_NUMBER_MODEL", raising=False)
    monkeypatch.delenv("KIMI_QUESTION_UNIT_MODEL", raising=False)
    monkeypatch.delenv("KIMI_MODEL", raising=False)
    monkeypatch.delenv("KIMI_TIMEOUT_S", raising=False)

    cfg = cfg_mod.load_kimi_config()

    assert cfg.number_model == "kimi-k2.6"
    assert cfg.model == "kimi-k2.6"


def test_sync_kimi_runtime_env_updates_effective_loaded_config(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_KIMI_CONFIG_PATH", raising=False)
    saved_env: dict[str, str] = {}

    def fake_set_user_environment_variable(name: str, value: str) -> None:
        text = (value or "").strip()
        if text:
            saved_env[name] = text
            monkeypatch.setenv(name, text)
        else:
            saved_env.pop(name, None)
            monkeypatch.delenv(name, raising=False)

    monkeypatch.setattr(cfg_mod, "set_user_environment_variable", fake_set_user_environment_variable)
    for name in (
        "KIMI_BASE_URL",
        "KIMI_API_KEY",
        "KIMI_QUESTION_NUMBER_MODEL",
        "KIMI_QUESTION_UNIT_MODEL",
        "KIMI_MODEL",
        "KIMI_TIMEOUT_S",
    ):
        monkeypatch.delenv(name, raising=False)

    path = cfg_mod._kimi_config_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {
                "base_url": "https://old.example/v1",
                "question_number_model": "old-number",
                "model": "old-unit",
                "timeout_s": 30.0,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    cfg_mod.sync_kimi_runtime_env(
        cfg_mod.KimiConfig(
            base_url="https://new.example/v1",
            api_key="kimi-key",
            number_model="kimi-number-model",
            model="kimi-unit-model",
            timeout_s=66.0,
        )
    )

    loaded = cfg_mod.load_kimi_config()
    assert saved_env["KIMI_BASE_URL"] == "https://new.example/v1"
    assert saved_env["KIMI_QUESTION_NUMBER_MODEL"] == "kimi-number-model"
    assert saved_env["KIMI_QUESTION_UNIT_MODEL"] == "kimi-unit-model"
    assert loaded.base_url == "https://new.example/v1"
    assert loaded.api_key == "kimi-key"
    assert loaded.number_model == "kimi-number-model"
    assert loaded.model == "kimi-unit-model"
    assert loaded.timeout_s == 66.0


def test_sync_qwen_runtime_env_updates_runtime_env_variables(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("APPDATA", str(tmp_path))
    monkeypatch.delenv("SJ_GENERATOR_QWEN_CONFIG_PATH", raising=False)
    saved_env: dict[str, str] = {}

    def fake_set_user_environment_variable(name: str, value: str) -> None:
        text = (value or "").strip()
        if text:
            saved_env[name] = text
            monkeypatch.setenv(name, text)
        else:
            saved_env.pop(name, None)
            monkeypatch.delenv(name, raising=False)

    monkeypatch.setattr(cfg_mod, "set_user_environment_variable", fake_set_user_environment_variable)
    for name in (
        "QWEN_BASE_URL",
        "QWEN_API_KEY",
        "QWEN_QUESTION_NUMBER_MODEL",
        "QWEN_QUESTION_UNIT_MODEL",
        "QWEN_MODEL",
        "QWEN_TIMEOUT_S",
        "QWEN_ACCOUNT_ACCESS_KEY_ID",
        "QWEN_ACCOUNT_ACCESS_KEY_SECRET",
        "ALIBABA_CLOUD_ACCESS_KEY_ID",
        "ALIBABA_CLOUD_ACCESS_KEY_SECRET",
    ):
        monkeypatch.delenv(name, raising=False)

    cfg_mod.sync_qwen_runtime_env(
        cfg_mod.QwenConfig(
            base_url="https://dashscope.example/v1",
            api_key="qwen-key",
            number_model="qwen-number-model",
            model="qwen-unit-model",
            account_access_key_id="akid",
            account_access_key_secret="aksecret",
            timeout_s=77.0,
        )
    )

    assert saved_env["QWEN_QUESTION_NUMBER_MODEL"] == "qwen-number-model"
    assert saved_env["QWEN_QUESTION_UNIT_MODEL"] == "qwen-unit-model"
    assert saved_env["ALIBABA_CLOUD_ACCESS_KEY_ID"] == "akid"
    assert saved_env["QWEN_BASE_URL"] == "https://dashscope.example/v1"
    assert saved_env["QWEN_API_KEY"] == "qwen-key"
    assert saved_env["QWEN_ACCOUNT_ACCESS_KEY_SECRET"] == "aksecret"


def test_pick_temperature_uses_kimi_k2_6_family_rule() -> None:
    assert _pick_temperature("kimi-k2.6") == 1.0
    assert _pick_temperature("kimi-k2.6-thinking") == 1.0
    assert _pick_temperature("deepseek-chat") == 0.0


def test_compare_highlight_marks_minority_model() -> None:
    got = compare_highlight_model_styles(
        model_sigs={"deepseek": '{"answer":"A"}', "kimi": '{"answer":"A"}', "qwen": '{"answer":"B"}'},
        round_no=1,
        round_matched_count=2,
    )
    assert got == {"qwen": "red"}


def test_compare_highlight_marks_all_when_first_round_failed() -> None:
    got = compare_highlight_model_styles(
        model_sigs={"deepseek": '{"answer":"A"}', "kimi": "", "qwen": '{"answer":"B"}'},
        round_no=1,
        round_matched_count=1,
    )
    assert got == {"deepseek": "red", "kimi": "yellow", "qwen": "red"}


def test_compare_highlight_skips_when_first_round_all_passed() -> None:
    got = compare_highlight_model_styles(
        model_sigs={"deepseek": '{"answer":"A"}', "kimi": '{"answer":"A"}', "qwen": '{"answer":"A"}'},
        round_no=1,
        round_matched_count=3,
    )
    assert got == {}


def test_compare_highlight_marks_non_first_round_minority_model() -> None:
    got = compare_highlight_model_styles(
        model_sigs={"deepseek": '{"answer":"A"}', "kimi": "", "qwen": '{"answer":"A"}'},
        round_no=2,
        round_matched_count=2,
    )
    assert got == {"kimi": "yellow"}


def test_compare_highlight_marks_all_non_empty_models_red_when_all_different() -> None:
    got = compare_highlight_model_styles(
        model_sigs={"deepseek": '{"answer":"A"}', "kimi": '{"answer":"B"}', "qwen": '{"answer":"C"}'},
        round_no=2,
        round_matched_count=1,
    )
    assert got == {"deepseek": "red", "kimi": "red", "qwen": "red"}


def test_compare_highlight_uses_same_fingerprint_semantics_as_verdict() -> None:
    deepseek = iq._normalize_question_obj_for_view(
        {"number": "1", "stem": "1. 题目一", "options": "A.甲\nB.乙", "answer": "a"}
    )
    kimi = iq._normalize_question_obj_for_view(
        {"number": "", "stem": "题目一", "options": "A.甲\nB.乙", "answer": "A"}
    )
    qwen = iq._normalize_question_obj_for_view(
        {"number": "3", "stem": "题目一", "options": "A. 丙\nB. 丁", "answer": "B"}
    )
    got = compare_highlight_model_styles(
        model_sigs={
            "deepseek": iq._fingerprint_question_obj(deepseek),
            "kimi": iq._fingerprint_question_obj(kimi),
            "qwen": iq._fingerprint_question_obj(qwen),
        },
        round_no=1,
        round_matched_count=2,
    )
    assert got == {"qwen": "red"}


def test_normalize_question_obj_for_view_splits_legacy_options_into_option_fields() -> None:
    got = iq._normalize_question_obj_for_view(
        {
            "question_type": "单选",
            "number": "2",
            "stem": "题目二",
            "options": "A.甲 B.乙 C.丙 D.丁",
            "answer": "D",
        }
    )
    assert "options" not in got
    assert got["option_1"] == "甲"
    assert got["option_2"] == "乙"
    assert got["option_3"] == "丙"
    assert got["option_4"] == "丁"


def test_fingerprint_question_obj_treats_option_fields_same_as_legacy_options() -> None:
    legacy = {
        "question_type": "单选",
        "number": "2",
        "stem": "题目二",
        "options": "A.甲\nB.乙\nC.丙\nD.丁",
        "answer": "D",
    }
    structured = {
        "question_type": "单选",
        "number": "2",
        "stem": "题目二",
        "option_1": "甲",
        "option_2": "乙",
        "option_3": "丙",
        "option_4": "丁",
        "answer": "D",
    }
    assert iq._fingerprint_question_obj(legacy) == iq._fingerprint_question_obj(structured)


def test_build_deepseek_balance_url_supports_common_base_url_forms() -> None:
    assert balance_mod._build_deepseek_balance_url("https://api.deepseek.com") == "https://api.deepseek.com/user/balance"
    assert balance_mod._build_deepseek_balance_url("https://api.deepseek.com/v1") == "https://api.deepseek.com/user/balance"
    assert (
        balance_mod._build_deepseek_balance_url("https://api.deepseek.com/v1/chat/completions")
        == "https://api.deepseek.com/user/balance"
    )


def test_format_deepseek_balance_infos_formats_currency_and_breakdown() -> None:
    got = balance_mod._format_deepseek_balance_infos(
        [
            {
                "currency": "CNY",
                "total_balance": "110",
                "granted_balance": "10",
                "topped_up_balance": "100",
            }
        ]
    )
    assert got == ["CNY ¥110.00（赠送 ¥10.00，充值 ¥100.00）"]


def test_build_kimi_balance_url_supports_common_base_url_forms() -> None:
    assert balance_mod._build_kimi_balance_url("https://api.moonshot.cn") == "https://api.moonshot.cn/v1/users/me/balance"
    assert balance_mod._build_kimi_balance_url("https://api.moonshot.cn/v1") == "https://api.moonshot.cn/v1/users/me/balance"
    assert (
        balance_mod._build_kimi_balance_url("https://api.moonshot.cn/v1/chat/completions")
        == "https://api.moonshot.cn/v1/users/me/balance"
    )


def test_describe_kimi_balance_payload_supports_direct_balance_field() -> None:
    got = balance_mod._describe_kimi_balance_payload({"balance": "15.5", "currency": "CNY"})
    assert got == "已配置，余额 CNY ¥15.50"


def test_describe_deepseek_balance_payload_formats_negative_balance() -> None:
    got = balance_mod._describe_deepseek_balance_payload(
        {
            "is_available": False,
            "balance_infos": [
                {
                    "currency": "CNY",
                    "total_balance": "-0.09",
                    "granted_balance": "0",
                    "topped_up_balance": "-0.09",
                }
            ],
        }
    )
    assert got == "已配置，余额 CNY ¥-0.09（赠送 ¥0.00，充值 ¥-0.09）（当前账户不可用）"


def test_describe_aliyun_account_balance_payload_supports_available_amount() -> None:
    got = balance_mod._describe_aliyun_account_balance_payload({"Data": {"AvailableAmount": "123.4", "Currency": "CNY"}})
    assert got == "已配置，阿里云账户余额 CNY ¥123.40"


def test_to_fast_timeout_cfg_caps_slow_balance_query_timeout() -> None:
    cfg = cfg_mod.KimiConfig(
        base_url="https://api.moonshot.cn/v1",
        api_key="test-key",
        number_model="kimi-k2.6",
        model="kimi-k2.6",
        timeout_s=120.0,
    )

    fast_cfg = cfg_mod.with_capped_timeout(cfg, 12.0)

    assert fast_cfg.timeout_s == 12.0
    assert fast_cfg.base_url == cfg.base_url
    assert fast_cfg.model == cfg.model


def test_run_checks_in_parallel_returns_on_first_failure() -> None:
    start = time.perf_counter()

    with pytest.raises(RuntimeError, match="boom"):
        run_callables_in_parallel_fail_fast(
            callables=[
                lambda: (_ for _ in ()).throw(RuntimeError("boom")),
                lambda: time.sleep(0.3),
            ],
            max_workers=2,
        )

    assert time.perf_counter() - start < 0.25


def test_run_analysis_tasks_supports_three_way_concurrency() -> None:
    tasks = [(i, f"题目{i}", "A") for i in range(5)]
    lock = threading.Lock()
    active = 0
    max_active = 0
    started: list[int] = []
    completed: list[int] = []

    def run_one(task: tuple[int, str, str]) -> int:
        nonlocal active, max_active
        with lock:
            active += 1
            max_active = max(max_active, active)
        time.sleep(0.05)
        with lock:
            active -= 1
        return task[0]

    run_tasks_in_parallel(
        tasks=tasks,
        max_workers=3,
        stop_cb=lambda: False,
        on_task_start=lambda current, total, task: started.append(current),
        on_task_done=lambda task, result: completed.append(result),
        on_task_failed=lambda task, exc: (_ for _ in ()).throw(exc),
        run_one=run_one,
    )

    assert started == [1, 2, 3, 4, 5]
    assert sorted(completed) == [0, 1, 2, 3, 4]
    assert max_active == 3


def test_normalize_ai_concurrency_keeps_allowed_values_and_falls_back() -> None:
    assert normalize_ai_concurrency(1) == 1
    assert normalize_ai_concurrency(2) == 2
    assert normalize_ai_concurrency(3) == 3
    assert normalize_ai_concurrency(4) == 4
    assert normalize_ai_concurrency(5) == 5
    assert normalize_ai_concurrency(6) == 6
    assert normalize_ai_concurrency(7) == 7
    assert normalize_ai_concurrency(8) == 8
    assert normalize_ai_concurrency(9) == 3
    assert normalize_ai_concurrency(None) == 3


def test_normalize_analysis_provider_keeps_allowed_values_and_falls_back() -> None:
    assert normalize_analysis_provider("deepseek") == "deepseek"
    assert normalize_analysis_provider("kimi") == "kimi"
    assert normalize_analysis_provider("qwen") == "qwen"
    assert normalize_analysis_provider("other") == "deepseek"
    assert normalize_analysis_provider(None) == "deepseek"


def test_normalize_analysis_model_name_keeps_input_and_falls_back() -> None:
    assert normalize_analysis_model_name("deepseek-reasoner") == "deepseek-reasoner"
    assert normalize_analysis_model_name("custom-model") == "custom-model"
    assert normalize_analysis_model_name("") == "deepseek-reasoner"
    assert normalize_analysis_model_name(None) == "deepseek-reasoner"


def test_normalize_default_repo_parent_dir_text_keeps_input_and_falls_back() -> None:
    assert normalize_default_repo_parent_dir_text("C:/repo-root") == "C:/repo-root"
    assert normalize_default_repo_parent_dir_text("") == str(default_repo_parent_dir())
    assert normalize_default_repo_parent_dir_text(None) == str(default_repo_parent_dir())


def test_process_excel_to_folder_mode_supports_legacy_headers(tmp_path) -> None:
    src = tmp_path / "1.1.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "questions"
    ws.append(["编号", "题目", "答案", "解析", "提取日期"])
    ws.append([1, "题目一\nA. 甲\nB. 乙", "A", "解析一", "2026-03-31"])
    wb.save(src)

    result = process_excel_to_folder_mode(src, export_date=date(2026, 3, 31))

    assert result.target_dir == tmp_path / "1.1"
    assert result.target_xlsx.exists()
    assert result.target_md.exists()

    saved = load_workbook(result.target_xlsx)
    ws2 = saved["questions"]
    assert [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))] == ["编号", "题目", "选项", "答案", "解析"]
    assert ws2.cell(2, 3).value in ("", None)


def test_process_source_files_to_folders_outputs_xlsx_and_md(monkeypatch, tmp_path) -> None:
    src = tmp_path / "资料一.txt"
    src.write_text("原始资料", encoding="utf-8")

    monkeypatch.setattr(batch_ai, "read_source_text", lambda path: "资料正文")
    monkeypatch.setattr(batch_ai, "generate_explanation", lambda client, inp: "解析一")
    monkeypatch.setattr(
        batch_ai,
        "import_questions_from_sources",
        lambda **kwargs: ImportResult(
            questions=[Question(number="1", stem="题目一", options="A.甲\nB.乙", answer="A", analysis="")],
            raw_items=[],
        ),
    )

    messages: list[str] = []
    results = batch_ai.process_source_files_to_folders(
        paths=[src],
        client=object(),
        kimi_client=object(),
        qwen_client=object(),
        export_date=date(2026, 3, 31),
        progress_cb=messages.append,
    )

    assert len(results) == 1
    result = results[0]
    assert result.target_dir == tmp_path / "资料一"
    assert result.target_xlsx.exists()
    assert result.target_md.exists()
    assert result.question_count == 1
    assert any("AI 解析准备中" in msg or "统计题数" in msg for msg in messages)
    saved = load_workbook(result.target_xlsx)
    ws = saved["questions"]
    assert ws.cell(2, 5).value == "解析一"

    progress_events: list[batch_ai.BatchAiProgress] = []
    batch_ai.process_source_files_to_folders(
        paths=[src],
        client=object(),
        kimi_client=object(),
        qwen_client=object(),
        export_date=date(2026, 3, 31),
        progress_info_cb=progress_events.append,
    )
    assert any(event.stage == "reading" for event in progress_events)
    assert any(event.stage == "generating_analysis" for event in progress_events)
    assert any(event.stage == "done" and event.question_count == 1 for event in progress_events)


def test_export_markdown_removes_extra_blank_line_for_convertible_multi_options() -> None:
    md = export_questions_to_markdown(
        excel_file_name="示例题库",
        export_date=date(2026, 3, 31),
        questions=[
            Question(
                number="1",
                stem="题目一",
                options="①. 甲\n②. 乙\n③. 丙\n④. 丁\n\nA. ①③\nB. ①②\nC. ②④\nD. ①④",
                answer="B",
                analysis="解析一",
            )
        ],
    )
    assert "④. 丁\n\nA. ①③" not in md
    assert "④. 丁\nA. ①③  B. ①②  C. ②④  D. ①④" in md


def test_export_markdown_orders_convertible_multi_statements_before_combo_lines() -> None:
    md = export_questions_to_markdown(
        excel_file_name="示例题库",
        export_date=date(2026, 3, 31),
        questions=[
            Question(
                number="1",
                stem="题目一",
                options="A. ①③\n①. 甲\nB. ①②\n②. 乙\n③. 丙\nD. ①④\n④. 丁\nC. ②④",
                answer="B",
                analysis="解析一",
            )
        ],
    )
    combo_line = "A. ①③  B. ①②  C. ②④  D. ①④"
    assert combo_line in md
    assert md.index("①. 甲") < md.index(combo_line)
    assert md.index("②. 乙") < md.index(combo_line)
    assert md.index("③. 丙") < md.index(combo_line)
    assert md.index("④. 丁") < md.index(combo_line)


def test_export_markdown_supports_convertible_multi_as_multi_mode() -> None:
    md = export_questions_to_markdown(
        excel_file_name="示例题库",
        export_date=date(2026, 3, 31),
        questions=[
            Question(
                number="1",
                stem="题目一",
                options="①. 甲\n②. 乙\n③. 丙\n④. 丁\nA. ①③\nB. ①②\nC. ②④\nD. ①④",
                answer="B",
                analysis="解析一",
            )
        ],
        convertible_multi_mode="as_multi",
    )
    assert "①. 甲" in md
    assert "②. 乙" in md
    assert "③. 丙" in md
    assert "④. 丁" in md
    assert "A. ①③" not in md
    assert "B. ①②" not in md


def test_process_source_files_to_folders_supports_controlled_concurrency(monkeypatch, tmp_path) -> None:
    src1 = tmp_path / "资料一.txt"
    src2 = tmp_path / "资料二.txt"
    src1.write_text("原始资料一", encoding="utf-8")
    src2.write_text("原始资料二", encoding="utf-8")

    monkeypatch.setattr(batch_ai, "read_source_text", lambda path: f"{path.stem} 正文")
    monkeypatch.setattr(
        batch_ai,
        "generate_explanation_result",
        lambda client, inp: explanations_mod.ExplanationResult(answer_text=inp.answer_text, analysis_text="解析一"),
    )
    monkeypatch.setattr(
        batch_ai,
        "import_questions_from_sources",
        lambda **kwargs: ImportResult(
            questions=[Question(number="1", stem="题目一", options="A.甲\nB.乙", answer="A", analysis="")],
            raw_items=[],
        ),
    )

    created: list[str] = []

    def make_client(tag: str):
        def factory():
            created.append(tag)
            return object()

        return factory

    results = batch_ai.process_source_files_to_folders(
        paths=[src1, src2],
        client_factory=make_client("deepseek"),
        analysis_client_factory=make_client("analysis"),
        kimi_client_factory=make_client("kimi"),
        qwen_client_factory=make_client("qwen"),
        max_workers=2,
        export_date=date(2026, 3, 31),
    )

    assert [item.target_dir.name for item in results] == ["资料一", "资料二"]
    assert len(created) == 8


def test_process_source_files_to_folders_supports_analysis_concurrency(monkeypatch, tmp_path) -> None:
    src = tmp_path / "资料一.txt"
    src.write_text("原始资料", encoding="utf-8")

    monkeypatch.setattr(batch_ai, "read_source_text", lambda path: "资料正文")
    monkeypatch.setattr(
        batch_ai,
        "import_questions_from_sources",
        lambda **kwargs: ImportResult(
            questions=[
                Question(number="1", stem="题目一", options="A.甲\nB.乙", answer="A", analysis=""),
                Question(number="2", stem="题目二", options="A.甲\nB.乙", answer="B", analysis=""),
            ],
            raw_items=[],
        ),
    )
    monkeypatch.setattr(
        batch_ai,
        "generate_explanation_result",
        lambda client, inp: explanations_mod.ExplanationResult(
            answer_text=inp.answer_text,
            analysis_text=f"解析-{inp.answer_text}",
        ),
    )

    created: list[str] = []

    def make_client():
        created.append("deepseek")
        return object()

    results = batch_ai.process_source_files_to_folders(
        paths=[src],
        client=object(),
        kimi_client=object(),
        qwen_client=object(),
        client_factory=make_client,
        max_analysis_workers=2,
        export_date=date(2026, 3, 31),
    )

    assert len(results) == 1
    saved = load_workbook(results[0].target_xlsx)
    ws = saved["questions"]
    assert ws.cell(2, 5).value == "解析-A"
    assert ws.cell(3, 5).value == "解析-B"
    assert len(created) == 2


def test_generate_explanation_result_fills_answer_when_missing() -> None:
    class FakeClient:
        def chat_text(self, *, system: str, user: str) -> str:
            assert "专业的教育助手" in system
            assert "题目文本" in user
            assert "如果“答案文本”为空" in user
            return "答案：a, c\n- A：**正确** 理由\n- B：**概念错误** 理由"

    result = explanations_mod.generate_explanation_result(
        FakeClient(),
        explanations_mod.ExplanationInputs(
            question_text="题目一\nA.甲\nB.乙\nC.丙\nD.丁",
            answer_text="",
        ),
    )

    assert result.answer_text == "AC"
    assert result.analysis_text == "A：**正确** 理由\nB：**概念错误** 理由"


def test_import_questions_from_sources_supports_question_level_concurrency(monkeypatch, tmp_path) -> None:
    src = tmp_path / "资料一.txt"
    src.write_text("原始资料", encoding="utf-8")

    monkeypatch.setattr(iq, "_get_question_number_list_verified", lambda **kwargs: ["1", "2", "3"])
    monkeypatch.setattr(
        iq,
        "_process_one_question",
        lambda **kwargs: (
            {"number": str(kwargs["index"]), "stem": f"题目{kwargs['index']}", "options": "", "answer": "A"},
            False,
            {},
        ),
    )

    created: list[str] = []

    def make_client(tag: str):
        def factory():
            created.append(tag)
            return object()

        return factory

    result = iq.import_questions_from_sources(
        client=object(),
        kimi_client=object(),
        qwen_client=object(),
        client_factory=make_client("deepseek"),
        kimi_client_factory=make_client("kimi"),
        qwen_client_factory=make_client("qwen"),
        sources=[(src, "资料正文")],
        strategy="per_question",
        max_question_workers=2,
    )

    assert [q.number for q in result.questions] == ["1", "2", "3"]
    assert len(created) == 9


def test_to_question_uses_original_analysis_when_present() -> None:
    q = iq._to_question(
        {
            "question_type": "单选",
            "number": "8",
            "stem": "题干",
            "option_1": "甲",
            "option_2": "乙",
            "option_3": "丙",
            "option_4": "丁",
            "answer": "A",
            "original_analysis": "原文自带解析",
        }
    )

    assert q.number == "8"
    assert q.analysis == "原文自带解析"


def test_import_prompt_templates_can_persist_and_render(monkeypatch, tmp_path) -> None:
    prompt_path = tmp_path / "import_prompts.json"
    monkeypatch.setenv("SJ_GENERATOR_IMPORT_PROMPTS_PATH", str(prompt_path))

    prompts = prompt_mod.default_import_prompts()
    prompts["question_number_list_system"] = "自定义题号规则"
    prompts["question_extract_user"] = "文件={{source_name}} 题号={{requested_number}} 文本={{chunk_text}}"
    prompt_mod.save_import_prompts(prompts)

    loaded = prompt_mod.load_import_prompts(force_reload=True)
    rendered = prompt_mod.render_import_prompt(
        "question_extract_user",
        source_name="资料.docx",
        requested_number="12",
        chunk_text="原文",
    )

    assert loaded["question_number_list_system"] == "自定义题号规则"
    assert rendered == "文件=资料.docx 题号=12 文本=原文"


def test_normalize_question_ref_list_supports_rich_json_and_legacy_strings() -> None:
    rich = iq._normalize_question_ref_list(
        [
            {"number": "1", "question_type": "单选"},
            {"number": "2", "question_type": "可转多选"},
            {"number": "2", "question_type": "多选"},
        ]
    )
    legacy = iq._normalize_question_ref_list(["3", "4", "4"])

    assert rich == [
        {"number": "1", "question_type": "单选"},
        {"number": "2", "question_type": "可转多选"},
    ]
    assert legacy == [{"number": "3"}, {"number": "4"}]


def test_normalize_question_ref_list_preserves_duplicate_warning() -> None:
    rich = iq._normalize_question_ref_list(
        [
            {"number": "7", "question_type": "单选", "duplicate_warning": "存在可疑的题号重复"},
            {"number": "7", "question_type": "单选"},
            {"number": "8", "question_type": "多选", "duplicate_warning": "duplicate suspected"},
        ]
    )

    assert rich == [
        {"number": "7", "question_type": "单选", "duplicate_warning": "存在可疑的题号重复"},
        {"number": "8", "question_type": "多选", "duplicate_warning": "存在可疑的题号重复"},
    ]


def test_parse_question_ref_response_text_supports_special_markers() -> None:
    with pytest.raises(iq._QuestionRefSpecialCaseError, match="题号重复"):
        iq._parse_question_ref_response_text("[题号重复]")

    with pytest.raises(iq._QuestionRefSpecialCaseError, match="所给文本无选择题目"):
        iq._parse_question_ref_response_text("[所给文本无选择题目]")
